from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class PerkUpdate:
    email: str
    is_weekend_ticket: bool
    food_voucher_count: int


def read_excel(path: str) -> list[PerkUpdate]:
    """Read perk updates from the first worksheet.

    Expected columns are:
    A email
    B 1 if weekend ticket, 0 for day ticket (or something that converts to boolean similarly)
    C number of food vouchers
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        perk_updates = []

        for row_number, values in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            if all(_is_blank(value) for value in values):
                continue

            perk_updates.append(
                PerkUpdate(
                    email=str(values[0]).strip(),
                    is_weekend_ticket=bool(values[1]),
                    food_voucher_count=int(values[2]),
                )
            )

        return perk_updates
    finally:
        workbook.close()


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""
