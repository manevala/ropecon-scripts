from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class TicketOrder:
    first_name: str
    last_name: str
    email: str
    language_code: str
    ticket_product_id: int
    number_of_tickets: int


def read_excel(path: str) -> list[TicketOrder]:
    """Read ticket orders from the first worksheet.

    Expected columns are:
    A first name, B last name, C email, D ticket product id,
    E language code, F number of tickets.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        ticket_orders = []

        for row_number, values in enumerate(
            worksheet.iter_rows(min_col=1, max_col=6, values_only=True),
            start=1,
        ):
            if all(_is_blank(value) for value in values):
                continue

            ticket_orders.append(
                TicketOrder(
                    first_name=str(values[0]).strip(),
                    last_name=str(values[1]).strip(),
                    email=str(values[2]).strip(),
                    language_code=str(values[4]).strip().lower(),
                    ticket_product_id=int(values[3]),
                    number_of_tickets=int(values[5]),
                )
            )

        return ticket_orders
    finally:
        workbook.close()


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""
